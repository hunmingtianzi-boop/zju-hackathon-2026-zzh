from __future__ import annotations

import argparse
import json
import os
import re
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable


SUPPORTED_EXTENSIONS = {".pdf", ".md", ".markdown", ".docx", ".xlsx", ".xlsm"}
LOADER_VERSION = "1.1.0"


@dataclass
class Section:
    title: str
    markdown: str
    source_path: str
    source_type: str
    book_id: str
    index: int
    level: int = 1
    page_start: int | None = None
    page_end: int | None = None
    sheet_name: str | None = None
    metadata: dict[str, Any] = field(default_factory=dict)


def clean_text(text: str) -> str:
    text = text.replace("\x08", "").replace("\ufffd", "")
    text = text.replace("\u3000", " ")
    text = re.sub(r"\.{4,}", "...", text)
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def slug_book_id(path: Path) -> str:
    return path.stem.strip() or "untitled"


def markdown_table(rows: list[list[Any]]) -> str:
    cleaned_rows = [
        [clean_text("" if cell is None else str(cell)).replace("|", r"\|").replace("\n", "<br>") for cell in row]
        for row in rows
    ]
    cleaned_rows = [row for row in cleaned_rows if any(cell for cell in row)]
    if not cleaned_rows:
        return ""

    width = max(len(row) for row in cleaned_rows)
    normalized = [row + [""] * (width - len(row)) for row in cleaned_rows]
    header = normalized[0]
    body = normalized[1:] or [[""] * width]
    lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join(["---"] * width) + " |",
    ]
    lines.extend("| " + " | ".join(row) + " |" for row in body)
    return "\n".join(lines)


def split_markdown_sections(
    markdown: str,
    source_path: str,
    source_type: str,
    book_id: str,
    default_title: str,
) -> list[Section]:
    heading_pattern = re.compile(r"^(#{1,6})\s+(.+?)\s*$")
    sections: list[Section] = []
    current_title = default_title
    current_level = 1
    current_lines: list[str] = []
    current_page: int | None = None
    section_start_page: int | None = None

    def flush() -> None:
        if not current_lines:
            return
        body = clean_text("\n".join(current_lines))
        meaningful_lines = [
            line for line in body.splitlines()
            if line.strip()
            and not re.match(r"^#{1,6}\s+", line.strip())
            and not re.match(r"^<!--.*-->$", line.strip())
        ]
        if not body or not meaningful_lines:
            return
        sections.append(
            Section(
                title=clean_text(current_title),
                markdown=body,
                source_path=source_path,
                source_type=source_type,
                book_id=book_id,
                index=len(sections),
                level=current_level,
                page_start=section_start_page,
                page_end=current_page,
            )
        )

    for raw_line in markdown.splitlines():
        line = raw_line.rstrip()
        page_match = re.match(r"^<!--\s*page:\s*(\d+)\s*-->\s*$", line)
        if page_match:
            current_page = int(page_match.group(1))
            if section_start_page is None:
                section_start_page = current_page
            current_lines.append(line)
            continue

        heading_match = heading_pattern.match(line)
        if heading_match:
            flush()
            current_title = heading_match.group(2)
            current_level = len(heading_match.group(1))
            current_lines = [line]
            section_start_page = current_page
        else:
            current_lines.append(line)

    flush()
    if sections:
        return sections

    fallback = clean_text(markdown)
    if not fallback:
        return []
    return [
        Section(
            title=default_title,
            markdown=fallback,
            source_path=source_path,
            source_type=source_type,
            book_id=book_id,
            index=0,
        )
    ]


class DocumentLoader:
    """Load textbooks and auxiliary materials into a Markdown-first Section model."""

    def load(self, path: str | os.PathLike[str]) -> list[Section]:
        path_obj = Path(path)
        if not path_obj.exists():
            raise FileNotFoundError(path_obj)
        if path_obj.is_dir():
            return self.load_many(path_obj)

        path_obj = path_obj.resolve()
        markdown = self.to_markdown(path_obj)
        source_type = path_obj.suffix.lower().lstrip(".")
        sections = split_markdown_sections(
            markdown=markdown,
            source_path=str(path_obj),
            source_type=source_type,
            book_id=slug_book_id(path_obj),
            default_title=path_obj.stem,
        )
        self._enrich_sections(sections, path_obj)
        if source_type in {"xlsx", "xlsm"}:
            for section in sections:
                section.sheet_name = section.title
                section.metadata["sheet_name"] = section.title
        return sections

    def load_many(self, paths: str | os.PathLike[str] | Iterable[str | os.PathLike[str]]) -> list[Section]:
        files: list[Path] = []
        if isinstance(paths, (str, os.PathLike)):
            root = Path(paths)
            if root.is_dir():
                files = sorted(
                    path for path in root.rglob("*")
                    if path.is_file() and path.suffix.lower() in SUPPORTED_EXTENSIONS
                )
            else:
                files = [root]
        else:
            for item in paths:
                path = Path(item)
                if path.is_dir():
                    files.extend(
                        sorted(
                            child for child in path.rglob("*")
                            if child.is_file() and child.suffix.lower() in SUPPORTED_EXTENSIONS
                        )
                    )
                else:
                    files.append(path)

        all_sections: list[Section] = []
        for file_path in files:
            all_sections.extend(self.load(file_path))
        return all_sections

    def to_markdown(self, path: str | os.PathLike[str]) -> str:
        path_obj = Path(path)
        ext = path_obj.suffix.lower()
        if ext in {".md", ".markdown"}:
            return path_obj.read_text(encoding="utf-8")
        if ext == ".pdf":
            return self._pdf_to_markdown(path_obj)
        if ext == ".docx":
            return self._docx_to_markdown(path_obj)
        if ext in {".xlsx", ".xlsm"}:
            return self._excel_to_markdown(path_obj)
        raise ValueError(f"Unsupported document type: {ext}. Supported: {sorted(SUPPORTED_EXTENSIONS)}")

    def write_markdown(self, source_path: str | os.PathLike[str], output_path: str | os.PathLike[str]) -> Path:
        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)
        output.write_text(self.to_markdown(source_path), encoding="utf-8")
        return output

    def build_manifest(self, sections: Iterable[Section]) -> dict[str, Any]:
        by_book: dict[str, dict[str, Any]] = {}
        section_list = list(sections)
        for section in section_list:
            record = by_book.setdefault(
                section.book_id,
                {
                    "book_id": section.book_id,
                    "source_path": section.source_path,
                    "source_type": section.source_type,
                    "sections": 0,
                    "chars": 0,
                    "pages": set(),
                },
            )
            record["sections"] += 1
            record["chars"] += len(section.markdown)
            if section.page_start is not None and section.page_end is not None:
                record["pages"].update(range(section.page_start, section.page_end + 1))

        books = []
        for record in by_book.values():
            pages = sorted(record.pop("pages"))
            record["page_count"] = len(pages)
            record["page_start"] = pages[0] if pages else None
            record["page_end"] = pages[-1] if pages else None
            books.append(record)

        return {
            "loader_version": LOADER_VERSION,
            "total_books": len(books),
            "total_sections": len(section_list),
            "total_chars": sum(len(section.markdown) for section in section_list),
            "books": sorted(books, key=lambda item: item["book_id"]),
        }

    def _pdf_to_markdown(self, path: Path) -> str:
        try:
            import fitz
        except ImportError as exc:
            raise ImportError("PDF loading requires PyMuPDF: pip install pymupdf") from exc

        doc = fitz.open(path)
        pages: list[str] = [f"# {path.stem}"]
        try:
            font_sizes = self._pdf_font_sizes(doc)
            common_size = sorted(font_sizes)[(len(font_sizes) - 1) // 2] if font_sizes else 0
            heading_sizes = sorted({size for size in font_sizes if size >= common_size + 1.0}, reverse=True)
            for page_index, page in enumerate(doc, start=1):
                page_lines = self._pdf_page_lines(page, common_size, heading_sizes)
                page_text = "\n\n".join(page_lines)
                if not page_text:
                    page_text = "> [Warning: No extractable text on this page.]"
                pages.append(f"<!-- page: {page_index} -->\n\n{page_text}")
        finally:
            doc.close()
        return "\n\n".join(pages)

    def _pdf_font_sizes(self, doc: Any) -> list[float]:
        sizes: list[float] = []
        sample_pages = min(len(doc), 40)
        for page in doc[:sample_pages]:
            data = page.get_text("dict")
            for block in data.get("blocks", []):
                if block.get("type") != 0:
                    continue
                for line in block.get("lines", []):
                    for span in line.get("spans", []):
                        text = clean_text(span.get("text", ""))
                        if text:
                            sizes.append(round(float(span.get("size", 0)), 1))
        return sizes

    def _pdf_page_lines(self, page: Any, common_size: float, heading_sizes: list[float]) -> list[str]:
        data = page.get_text("dict")
        blocks = data.get("blocks", [])
        blocks.sort(key=lambda block: (block.get("bbox", [0, 0])[1], block.get("bbox", [0, 0])[0]))

        page_height = float(page.rect.height)
        page_lines: list[str] = []
        for block in blocks:
            if block.get("type") != 0:
                continue
            bbox = block.get("bbox", [0, 0, 0, 0])
            block_lines: list[str] = []
            max_size = 0.0
            bold = False
            for line in block.get("lines", []):
                text_parts: list[str] = []
                for span in line.get("spans", []):
                    text = clean_text(span.get("text", ""))
                    if not text:
                        continue
                    text_parts.append(text)
                    max_size = max(max_size, round(float(span.get("size", 0)), 1))
                    font_name = str(span.get("font", "")).lower()
                    flags = int(span.get("flags", 0))
                    bold = bold or "bold" in font_name or bool(flags & 16)
                line_text = clean_text(" ".join(text_parts))
                if line_text:
                    block_lines.append(line_text)

            text = clean_text(" ".join(block_lines))
            if not text:
                continue
            if (bbox[1] < 45 or bbox[3] > page_height - 45) and len(text) < 24:
                continue
            page_lines.append(self._format_pdf_block(text, max_size, common_size, heading_sizes, bold))
        return page_lines

    def _format_pdf_block(
        self,
        text: str,
        size: float,
        common_size: float,
        heading_sizes: list[float],
        bold: bool,
    ) -> str:
        level = self._heading_level(text, size, common_size, heading_sizes, bold)
        return f"{'#' * level} {text}" if level else text

    def _heading_level(
        self,
        text: str,
        size: float,
        common_size: float,
        heading_sizes: list[float],
        bold: bool,
    ) -> int:
        normalized = clean_text(text)
        if re.match(r"^(绪论|前言|导论|总论)$", normalized):
            return 1
        if re.match(r"^(chapter|part|section)\s+\d+", normalized, flags=re.IGNORECASE):
            return 1
        if re.match(r"^第\s*[一二三四五六七八九十百零〇\d]+\s*[篇编章]\b", normalized):
            return 1
        if re.match(r"^第\s*[一二三四五六七八九十百零〇\d]+\s*节\b", normalized):
            return 2
        if re.match(r"^[一二三四五六七八九十]+[、.．]\s*", normalized):
            return 3
        if re.match(r"^[(（][一二三四五六七八九十\d]+[)）]\s*", normalized):
            return 4
        if re.match(r"^\d+(?:\.\d+){1,3}\s+", normalized):
            return 3
        if len(normalized) > 60 or common_size <= 0:
            return 0
        if size >= common_size + 2.5:
            return 1 if not heading_sizes or size >= heading_sizes[0] - 0.2 else 2
        if size >= common_size + 1.2 or (bold and size >= common_size + 0.5):
            return 3
        return 0

    def _docx_to_markdown(self, path: Path) -> str:
        try:
            from docx import Document
            from docx.table import Table
            from docx.text.paragraph import Paragraph
            from docx.oxml.table import CT_Tbl
            from docx.oxml.text.paragraph import CT_P
        except ImportError as exc:
            raise ImportError("Word loading requires python-docx: pip install python-docx") from exc

        document = Document(path)
        lines: list[str] = [f"# {path.stem}"]

        for child in document.element.body.iterchildren():
            if isinstance(child, CT_P):
                paragraph = Paragraph(child, document)
                text = clean_text(paragraph.text)
                if not text:
                    continue
                style_name = paragraph.style.name if paragraph.style else ""
                heading_match = re.match(r"(?:Heading|标题)\s*([1-6])", style_name, flags=re.IGNORECASE)
                if heading_match:
                    level = int(heading_match.group(1))
                    lines.append(f"{'#' * level} {text}")
                else:
                    lines.append(text)
            elif isinstance(child, CT_Tbl):
                table = Table(child, document)
                rows = [[cell.text for cell in row.cells] for row in table.rows]
                table_md = markdown_table(rows)
                if table_md:
                    lines.append(table_md)

        return "\n\n".join(lines)

    def _enrich_sections(self, sections: list[Section], path: Path) -> None:
        try:
            stat = path.stat()
            size = stat.st_size
            mtime = stat.st_mtime
        except OSError:
            size = 0
            mtime = None
        for index, section in enumerate(sections):
            section.index = index
            section.metadata.update(
                {
                    "loader_version": LOADER_VERSION,
                    "source_name": path.name,
                    "source_size": str(size),
                    "source_mtime": str(mtime) if mtime is not None else "",
                    "char_count": str(len(section.markdown)),
                    "section_id": f"{section.book_id}:{index:04d}",
                }
            )

    def _excel_to_markdown(self, path: Path) -> str:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ImportError("Excel loading requires openpyxl: pip install openpyxl") from exc

        workbook = load_workbook(path, read_only=True, data_only=True)
        lines: list[str] = [f"# {path.stem}"]
        try:
            for sheet in workbook.worksheets:
                rows = [list(row) for row in sheet.iter_rows(values_only=True)]
                table_md = markdown_table(rows)
                if not table_md:
                    continue
                lines.append(f"## {sheet.title}")
                lines.append(table_md)
        finally:
            workbook.close()
        return "\n\n".join(lines)


def main() -> None:
    parser = argparse.ArgumentParser(description="Load PDF/Markdown/Word/Excel sources into Markdown sections.")
    parser.add_argument("paths", nargs="+", help="Files or folders to load.")
    parser.add_argument("--export-md", help="Directory for exported Markdown files.")
    parser.add_argument("--manifest", help="Write a JSON loading manifest.")
    args = parser.parse_args()

    loader = DocumentLoader()
    sections = loader.load_many(args.paths)

    if args.export_md:
        output_dir = Path(args.export_md)
        output_dir.mkdir(parents=True, exist_ok=True)
        for input_path in args.paths:
            path = Path(input_path)
            candidates = (
                sorted(child for child in path.rglob("*") if child.is_file() and child.suffix.lower() in SUPPORTED_EXTENSIONS)
                if path.is_dir()
                else [path]
            )
            for candidate in candidates:
                loader.write_markdown(candidate, output_dir / f"{candidate.stem}.md")

    if args.manifest:
        manifest_path = Path(args.manifest)
        manifest_path.parent.mkdir(parents=True, exist_ok=True)
        manifest_path.write_text(
            json.dumps(loader.build_manifest(sections), ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    print(f"Loaded {len(sections)} sections from {len(args.paths)} input path(s).")
    by_book: dict[str, int] = {}
    for section in sections:
        by_book[section.book_id] = by_book.get(section.book_id, 0) + 1
    for book_id, count in sorted(by_book.items()):
        print(f"- {book_id}: {count} sections")


if __name__ == "__main__":
    main()
